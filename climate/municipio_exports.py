# climate/municipio_exports.py
"""
Exportação dos indicadores climáticos de um MUNICÍPIO (não fazenda —
ver farms/exports.py pro equivalente do painel privado) pra um único
.xlsx multi-aba, PÚBLICO, sem login — pra pesquisadores/gestores
reusarem o dado fora da plataforma (fora do roadmap original, pedido
do usuário depois da home pública). Mesma técnica de farms/exports.py:
`openpyxl` já é dependência desde a Etapa 6, zero lib nova.

Reaproveita climate.municipio_indicators e climate.trends inteiros —
nenhum cálculo novo aqui, só serialização em planilha. Município sem
CHIRPS é responsabilidade de quem chama (api/views.py checa antes de
gerar o workbook, pra nunca devolver um .xlsx vazio/quebrado).
"""
from django.utils import timezone
from openpyxl import Workbook

from climate import trends
from climate.models import ChirpsData

from . import municipio_indicators as mi

# Citação recomendada do CHIRPS — texto fixo, pedido explícito do
# usuário (rigor científico: quem reusa o dado deve citar a fonte
# original, não a plataforma).
CITACAO_CHIRPS = (
    "Funk, C., Peterson, P., Landsfeld, M. et al. The climate hazards "
    "infrared precipitation with stations—a new environmental "
    "record for monitoring extremes. Sci Data 2, 150066 (2015). "
    "https://doi.org/10.1038/sdata.2015.66"
)

ESCALAS_SPI_EXPORT = (1, 3, 6, 12)

MESES_NOME = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho",
    "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def gerar_workbook_municipio(municipio):
    """Monta o .xlsx completo — 7 abas, ordem: metadados/citação
    primeiro, resumo do estado atual, depois dado cru, depois
    indicadores derivados, tendências por último."""
    workbook = Workbook()
    workbook.remove(workbook.active)  # aba padrão vazia do Workbook()

    _aba_metadados(workbook, municipio)
    _aba_resumo(workbook, municipio)
    _aba_precipitacao_diaria(workbook, municipio)
    _aba_spi(workbook, municipio)
    _aba_climatologia_mensal(workbook, municipio)
    _aba_indicadores_anuais(workbook, municipio)
    _aba_tendencias(workbook, municipio)

    return workbook


def _nova_aba(workbook, titulo, cabecalho):
    aba = workbook.create_sheet(title=titulo)
    aba.append(cabecalho)
    return aba


def _aba_metadados(workbook, municipio):
    aba = _nova_aba(workbook, "Metadados", ["Campo", "Valor"])
    centroide = municipio.geom.centroid  # x = longitude, y = latitude

    primeira = ChirpsData.objects.filter(municipio=municipio).order_by("date").first()
    ultima = ChirpsData.objects.filter(municipio=municipio).order_by("-date").first()
    periodo = f"{primeira.date.isoformat()} a {ultima.date.isoformat()}" if primeira and ultima else "—"

    aba.append(["Fonte dos dados de precipitação", "CHIRPS (Climate Hazards Group InfraRed Precipitation with Station data)"])
    aba.append(["Resolução espacial", "~0.05° (~5 km)"])
    aba.append(["Cobertura temporal da fonte", "Diária, desde 01/01/1981"])
    aba.append(["Método de extração", "Google Earth Engine (coleção UCSB-CHG/CHIRPS/DAILY), média zonal por município"])
    aba.append(["Município", municipio.nome])
    aba.append(["UF", municipio.uf])
    aba.append(["Código IBGE", municipio.codigo_ibge])
    aba.append(["Latitude do centroide", centroide.y])
    aba.append(["Longitude do centroide", centroide.x])
    aba.append(["Período dos dados neste arquivo", periodo])
    aba.append(["Data de extração deste arquivo", timezone.localtime().replace(tzinfo=None).strftime("%d/%m/%Y %H:%M")])
    aba.append(["Citação recomendada do CHIRPS", CITACAO_CHIRPS])
    aba.append(["Plataforma de origem", "MonitorChuva MT — monitoramento de precipitação e índice de seca em Mato Grosso"])


def _aba_resumo(workbook, municipio):
    """Snapshot do que a home pública mostra HOJE — os cards de FASE 1
    e FASE 2, sem recalcular nada, só formatados como linhas."""
    aba = _nova_aba(workbook, "Resumo (Indicadores Atuais)", ["Indicador", "Valor"])

    for escala in ESCALAS_SPI_EXPORT:
        atual = mi.spi_atual(municipio, escala)
        rotulo = f"SPI-{escala} atual ({atual['date'].strftime('%m/%Y')})" if atual else f"SPI-{escala} atual"
        valor = f"{atual['value']:.2f} ({atual['classification']})" if atual else "—"
        aba.append([rotulo, valor])

    anomalia = mi.anomalia_mensal(municipio)
    if anomalia:
        aba.append([
            f"Anomalia de chuva ({anomalia['mes']:02d}/{anomalia['ano']})",
            f"{anomalia['chuva_mm']:.1f} mm (média histórica: {anomalia['media_historica_mm']:.1f} mm, "
            f"{anomalia['anomalia_percentual']:.0f}% da média)",
        ])
    else:
        aba.append(["Anomalia de chuva", "—"])

    percentil = mi.percentil_historico_mensal(municipio)
    if percentil:
        aba.append([
            f"Percentil histórico ({percentil['mes']:02d}/{percentil['ano']})",
            f"Percentil {percentil['percentil']:.0f} — {percentil['posicao_mais_seco']}º mais seco desde {percentil['primeiro_ano_historico']}",
        ])
    else:
        aba.append(["Percentil histórico", "—"])

    for item in mi.acumulados_municipio(municipio):
        valor = f"{item['chirps_mm']:.1f} mm" if item["chirps_mm"] is not None else "—"
        aba.append([f"Acumulado {item['dias']} dias", valor])

    veranico = mi.veranico(municipio)
    if veranico:
        recente, recorde = veranico["recente_12_meses"], veranico["recorde_historico"]
        aba.append([
            "Dias secos consecutivos — últimos 12 meses",
            f"{recente['dias']} dias ({recente['inicio']} a {recente['fim']})" if recente else "—",
        ])
        aba.append([
            "Dias secos consecutivos — recorde histórico",
            f"{recorde['dias']} dias ({recorde['inicio']} a {recorde['fim']})" if recorde else "—",
        ])
    else:
        aba.append(["Dias secos consecutivos", "—"])

    dc = mi.dias_chuvosos(municipio)
    if dc:
        aba.append([
            f"Dias chuvosos ({dc['ano']})",
            f">1mm: {dc['limiares']['1']} | >5mm: {dc['limiares']['5']} | >10mm: {dc['limiares']['10']}",
        ])
    else:
        aba.append(["Dias chuvosos", "—"])

    ic = mi.intensidade_chuva(municipio)
    if ic and ic["intensidade_mm_por_dia_chuvoso"] is not None:
        aba.append([f"Intensidade da chuva ({ic['ano']})", f"{ic['intensidade_mm_por_dia_chuvoso']:.2f} mm/dia chuvoso"])
    else:
        aba.append(["Intensidade da chuva", "—"])

    recordes = mi.recordes(municipio)
    if recordes:
        aba.append(["Ano mais chuvoso", f"{recordes['ano_mais_chuvoso']['ano']} ({recordes['ano_mais_chuvoso']['total_mm']:.0f} mm)"])
        aba.append(["Ano mais seco", f"{recordes['ano_mais_seco']['ano']} ({recordes['ano_mais_seco']['total_mm']:.0f} mm)"])
        aba.append(["Mês mais chuvoso já registrado", f"{recordes['mes_mais_chuvoso']['date'].strftime('%m/%Y')} ({recordes['mes_mais_chuvoso']['total_mm']:.0f} mm)"])
        aba.append(["Mês mais seco já registrado", f"{recordes['mes_mais_seco']['date'].strftime('%m/%Y')} ({recordes['mes_mais_seco']['total_mm']:.1f} mm)"])
    else:
        aba.append(["Recordes", "—"])


def _aba_precipitacao_diaria(workbook, municipio):
    """Série bruta completa — a base de tudo, pra quem quiser reproduzir
    qualquer gráfico/indicador da ferramenta do zero. Mesmo padrão de
    farms/exports.py:_aba_chirps (dezenas de milhares de linhas, de
    propósito — dado cru, não resumo)."""
    aba = _nova_aba(workbook, "Precipitação Diária (CHIRPS)", ["Data", "Precipitação (mm)"])
    for registro in ChirpsData.objects.filter(municipio=municipio).order_by("date"):
        aba.append([registro.date, registro.value])


def _aba_spi(workbook, municipio):
    aba = _nova_aba(workbook, "SPI", ["Data", "Escala", "Valor", "Classificação"])
    for escala in ESCALAS_SPI_EXPORT:
        for ponto in mi.spi_serie(municipio, escala):
            aba.append([ponto["date"], f"SPI-{escala}", round(ponto["value"], 2), ponto["classification"]])


def _aba_climatologia_mensal(workbook, municipio):
    aba = _nova_aba(workbook, "Climatologia Mensal", [
        "Mês", "Média (mm)", "Mediana (mm)", "P25 (mm)", "P75 (mm)", "Mínimo (mm)", "Máximo (mm)", "N Anos",
    ])
    climatologia = mi.climatologia_mensal(municipio)
    for mes in sorted(climatologia):
        c = climatologia[mes]
        aba.append([
            MESES_NOME[mes - 1], round(c["media"], 1), round(c["mediana"], 1), round(c["p25"], 1),
            round(c["p75"], 1), round(c["minimo"], 1), round(c["maximo"], 1), c["n_anos"],
        ])


def _aba_indicadores_anuais(workbook, municipio):
    """Uma linha por ano — junta 4 indicadores anuais diferentes (total,
    dias chuvosos nos 3 limiares, intensidade, veranico máximo) que na
    home aparecem em gráficos separados, mas aqui cabem juntos numa
    tabela só, pra análise em planilha/R/pandas."""
    aba = _nova_aba(workbook, "Indicadores Anuais", [
        "Ano", "Total Anual (mm)", "Dias Chuvosos (>1mm)", "Dias Chuvosos (>5mm)", "Dias Chuvosos (>10mm)",
        "Intensidade (mm/dia chuvoso)", "Veranico Máximo (dias)", "Veranico Máximo — Início", "Veranico Máximo — Fim",
    ])

    totais = trends.totais_anuais(municipio)
    dias_chuvosos = mi.dias_chuvosos_serie_anual(municipio) or {}
    intensidade = (mi.intensidade_serie_anual(municipio) or {}).get("serie", {})
    veranico = (mi.veranico_maximo_serie_anual(municipio) or {}).get("serie", {})

    for ano in sorted(totais):
        dias1 = dias_chuvosos.get("1", {}).get("serie", {}).get(ano)
        dias5 = dias_chuvosos.get("5", {}).get("serie", {}).get(ano)
        dias10 = dias_chuvosos.get("10", {}).get("serie", {}).get(ano)
        intensidade_ano = intensidade.get(ano, {}).get("intensidade_mm_por_dia_chuvoso")
        v = veranico.get(ano)
        aba.append([
            ano, round(totais[ano], 1), dias1, dias5, dias10,
            round(intensidade_ano, 2) if intensidade_ano is not None else None,
            v["dias"] if v else None, v["inicio"] if v else None, v["fim"] if v else None,
        ])


def _aba_tendencias(workbook, municipio):
    """Resultado COMPLETO do teste (não só o slope) — S, Z, p-valor,
    significância — pra ser verificável, mais a interpretação em texto
    (mesma regra de honestidade dos gráficos da home, via
    mi.interpretar_tendencia — não duplicada, importada)."""
    aba = _nova_aba(workbook, "Tendências (Mann-Kendall)", [
        "Série", "Slope (Sen's)", "Unidade", "Intercepto", "Mann-Kendall S", "Z", "p-valor",
        "Significativo (p<0.05)", "Direção", "N Anos", "Interpretação",
    ])

    dias_chuvosos_1mm = (mi.dias_chuvosos_serie_anual(municipio) or {}).get("1", {})
    intensidade = mi.intensidade_serie_anual(municipio) or {}
    veranico = mi.veranico_maximo_serie_anual(municipio) or {}

    linhas = [
        ("Total Anual de Chuva", mi.tendencia_mann_kendall(municipio), "mm/ano", "aumento", "redução", False),
        ("Dias Chuvosos (>1mm)", dias_chuvosos_1mm.get("tendencia"), "dias/ano", "aumento", "redução", False),
        ("Intensidade da Chuva", intensidade.get("tendencia"), "mm/dia chuvoso/ano", "intensificação", "redução de intensidade", True),
        ("Veranico Máximo (dias secos consecutivos)", veranico.get("tendencia"), "dias/ano", "aumento", "redução", False),
    ]

    for nome, t, unidade, rotulo_aumento, rotulo_reducao, estavel in linhas:
        interpretacao = mi.interpretar_tendencia(t, rotulo_aumento, rotulo_reducao, estavel)
        if t is None:
            aba.append([nome, None, unidade, None, None, None, None, None, None, None, interpretacao])
            continue
        aba.append([
            nome, round(t["slope_mm_por_ano"], 4), unidade, round(t["intercepto_sen"], 4),
            t["mann_kendall_s"], round(t["mann_kendall_z"], 3), round(t["p_valor"], 4),
            "Sim" if t["significativo"] else "Não", t["direcao"], t["n_anos"], interpretacao,
        ])
