# farms/exports.py
"""
Exportação de todo o dado de uma fazenda pra um único .xlsx com várias
abas (Etapa 11) — pra abrir em Excel, R, Python, SPSS etc. fora da
plataforma. `openpyxl` já é dependência do projeto desde a Etapa 6
(importação de planilha), sem lib nova.

Cada aba é uma tabela "crua" (uma linha por registro, sem soma nem
formatação de card) — é dado pra reanalisar em outro lugar, não um
resumo visual (isso é o relatório de impressão, `farms/views.py:
relatorio_fazenda`).
"""
from django.utils import timezone
from openpyxl import Workbook

from alerts.models import Alert
from climate.models import ChirpsData, ChirpsValidation, Projection, RainfallData
from spi.models import SpiResult


def gerar_workbook_fazenda(fazenda):
    workbook = Workbook()
    workbook.remove(workbook.active)  # aba padrão vazia do Workbook() — cada aba abaixo é criada explicitamente

    _aba_fazenda(workbook, fazenda)
    _aba_estacoes(workbook, fazenda)
    _aba_talhoes(workbook, fazenda)
    _aba_chuva_local(workbook, fazenda)
    _aba_chirps(workbook, fazenda)
    _aba_spi(workbook, fazenda)
    _aba_validacao_chirps(workbook, fazenda)
    _aba_alertas(workbook, fazenda)
    _aba_cenarios_futuros(workbook, fazenda)

    return workbook


def _nova_aba(workbook, titulo, cabecalho):
    aba = workbook.create_sheet(title=titulo)
    aba.append(cabecalho)
    return aba


def _aba_fazenda(workbook, fazenda):
    aba = _nova_aba(workbook, "Fazenda", ["Campo", "Valor"])
    aba.append(["Nome", fazenda.name])
    aba.append(["Município", f"{fazenda.municipio.nome}/{fazenda.municipio.uf}"])
    aba.append(["Área (ha)", fazenda.area])
    aba.append(["Cultura", fazenda.crop])
    aba.append(["Latitude", fazenda.latitude])
    aba.append(["Longitude", fazenda.longitude])
    aba.append(["Observações", fazenda.notes])


def _aba_estacoes(workbook, fazenda):
    aba = _nova_aba(workbook, "Estações", ["Nome", "Tipo", "Latitude", "Longitude"])
    for estacao in fazenda.stations.all():
        aba.append([estacao.name, estacao.get_station_type_display(), estacao.latitude, estacao.longitude])


def _aba_talhoes(workbook, fazenda):
    aba = _nova_aba(workbook, "Talhões", ["Nome", "Cultura", "Área (ha)", "Latitude", "Longitude"])
    for talhao in fazenda.talhoes.all():
        aba.append([talhao.name, talhao.crop, talhao.area, talhao.latitude, talhao.longitude])


def _aba_chuva_local(workbook, fazenda):
    aba = _nova_aba(workbook, "Chuva Local", ["Data", "Horário", "Valor (mm)", "Origem", "Estação", "Observações"])
    registros = (
        RainfallData.objects.filter(farm=fazenda).exclude(source_type="chirps")
        .select_related("station").order_by("date")
    )
    for registro in registros:
        aba.append([
            registro.date, registro.time, registro.value, registro.get_source_type_display(),
            registro.station.name, registro.notes,
        ])


def _aba_chirps(workbook, fazenda):
    """
    Série completa do CHIRPS do MUNICÍPIO da fazenda (não filtrada por
    estação — é a mesma média zonal pro município inteiro, ver
    docs/DECISOES.md sobre a Etapa 3.1). Pode ter dezenas de milhares
    de linhas (histórico desde 1981); é dado bruto de propósito, pra
    quem for reanalisar a série inteira.
    """
    aba = _nova_aba(workbook, "CHIRPS (Município)", ["Data", "Valor (mm)"])
    registros = ChirpsData.objects.filter(municipio=fazenda.municipio).order_by("date")
    for registro in registros:
        aba.append([registro.date, registro.value])


def _aba_spi(workbook, fazenda):
    aba = _nova_aba(workbook, "SPI", ["Data", "Escala", "Valor", "Classificação", "Estação"])
    registros = SpiResult.objects.filter(farm=fazenda).select_related("station").order_by("station", "scale", "date")
    for registro in registros:
        aba.append([
            registro.date, f"SPI-{registro.scale}", registro.value,
            registro.get_classification_display(), registro.station.name,
        ])


def _aba_validacao_chirps(workbook, fazenda):
    aba = _nova_aba(workbook, "Validação CHIRPS", [
        "Estação", "Nº Pares", "R²", "RMSE (mm)", "MAE (mm)", "MBE (mm)",
        "Índice d", "Índice c", "Desempenho", "Calculado em",
    ])
    registros = ChirpsValidation.objects.filter(farm=fazenda).select_related("station")
    for registro in registros:
        aba.append([
            registro.station.name, registro.n_pares, registro.r2, registro.rmse, registro.mae, registro.mbe,
            registro.indice_d, registro.indice_c, registro.get_desempenho_c_display(),
            timezone.localtime(registro.calculado_em).replace(tzinfo=None) if registro.calculado_em else None,
        ])


def _aba_alertas(workbook, fazenda):
    aba = _nova_aba(workbook, "Alertas", ["Tipo", "Mensagem", "Estação", "Ativo", "Criado em"])
    registros = Alert.objects.filter(farm=fazenda).select_related("station").order_by("-created_at")
    for registro in registros:
        aba.append([
            registro.get_alert_type_display(), registro.message,
            registro.station.name if registro.station else "",
            "Sim" if registro.is_active else "Não",
            timezone.localtime(registro.created_at).replace(tzinfo=None),
        ])


def _aba_cenarios_futuros(workbook, fazenda):
    aba = _nova_aba(workbook, "Cenários Futuros", ["Mês", "Cenário", "Valor (mm)", "Estação"])
    registros = Projection.objects.filter(farm=fazenda).select_related("station").order_by("date", "scenario")
    for registro in registros:
        aba.append([registro.date, registro.scenario, registro.value, registro.station.name if registro.station else ""])
